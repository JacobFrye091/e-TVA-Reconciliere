"""Excel report: summary with suggestions + detailed differences."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_RED = PatternFill("solid", start_color="FFC7CE")
_BOLD = Font(bold=True)
_ITALIC = Font(italic=True)
_NOTA = ("Notă: valorile sugerate sunt informative. Platforma nu se implică în corecții — "
        "corectarea jurnalului contabil rămâne responsabilitatea internă a firmei/contabilului.")

_SUMAR_HEADER = ["Categorie", "Baza firma", "TVA firma", "Baza ANAF",
                 "TVA ANAF", "Baza sugerata", "TVA sugerata", "Status"]
_DIFF_HEADER = ["Tip diferenta", "CUI partener", "Nr factura", "Categorie",
                "Baza firma", "TVA firma", "Baza ANAF", "TVA ANAF",
                "Delta baza", "Delta TVA"]
_SUSPECTE_HEADER = ["Document", "Data", "CUI partener", "Baza", "TVA",
                    "Linie D300", "De ce"]
_NEELUCIDATE_HEADER = ["Linie", "Denumire", "Delta baza", "Delta TVA",
                       "Explicatie"]
_MOTIV_SUSPECT = {
    "candidat": "Posibila cauza — suma explica exact diferenta",
    "aproximativ": "Cea mai apropiata suma gasita (nu explica exact diferenta)",
}


def write_report(result, suggestions, path, client_name, period) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sumar"
    ws["A1"] = f"Client: {client_name}"
    ws["A2"] = f"Perioada: {period}"
    ws["A1"].font = ws["A2"].font = _BOLD
    ws.append([_NOTA])
    ws["A3"].font = _ITALIC
    ws.append(_SUMAR_HEADER)
    for cell in ws[4]:
        cell.font = _BOLD
    for s in suggestions:
        ws.append([s["category"], s["company_base"], s["company_vat"],
                   s["anaf_base"], s["anaf_vat"], s["suggested_base"],
                   s["suggested_vat"], s["status"]])
        if s["status"] == "de_verificat":
            for cell in ws[ws.max_row]:
                cell.fill = _RED

    wd = wb.create_sheet("Diferente")
    wd.append(_DIFF_HEADER)
    for cell in wd[1]:
        cell.font = _BOLD
    for d in result.differences:
        c, a = d["company"], d["anaf"]
        wd.append([d["diff_type"], d["partner_cui"], d["invoice_no"],
                   d["category"],
                   c["base"] if c else "", c["vat"] if c else "",
                   a["base"] if a else "", a["vat"] if a else "",
                   d["delta_base"], d["delta_vat"]])
    wb.save(path)


_SUMAR_HEADER_LINII = ["Linie D300", "Denumire", "Baza firma", "TVA firma",
                       "Baza ANAF", "TVA ANAF", "Baza sugerata",
                       "TVA sugerata", "Status"]
_DIFF_HEADER_LINII = ["Tip diferenta", "Linie D300", "Denumire",
                      "Baza firma", "TVA firma", "Baza ANAF", "TVA ANAF",
                      "Delta baza", "Delta TVA"]


def write_report_lines(result, suggestions, path, client_name, period,
                       facturi_suspecte=None, linii_neelucidate=None) -> None:
    """Same report layout as write_report, but for D300-line-level results
    (real ANAF e-TVA precompleted document — no invoice detail).

    `facturi_suspecte`/`linii_neelucidate` (optionale, vezi
    portal/app.py::_facturi_suspecte) adauga foaia "Facturi de verificat",
    asezata PRIMA fiindca e raspunsul la intrebarea cu care incepe orice
    verificare: de unde incep. Fara ea, raportul dus la dosar ar arata doar
    liniile cu diferente, iar contabilul ar trebui sa revina in aplicatie ca
    sa afle care facturi sunt de fapt de verificat."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sumar"
    ws["A1"] = f"Client: {client_name}"
    ws["A2"] = f"Perioada: {period}"
    ws["A1"].font = ws["A2"].font = _BOLD
    ws.append([_NOTA])
    ws["A3"].font = _ITALIC
    ws.append(_SUMAR_HEADER_LINII)
    for cell in ws[4]:
        cell.font = _BOLD
    for s in suggestions:
        ws.append([s["line_no"], s["label"], s["company_base"], s["company_vat"],
                   s["anaf_base"], s["anaf_vat"], s["suggested_base"],
                   s["suggested_vat"], s["status"]])
        if s["status"] == "de_verificat":
            for cell in ws[ws.max_row]:
                cell.fill = _RED

    wd = wb.create_sheet("Diferente")
    wd.append(_DIFF_HEADER_LINII)
    for cell in wd[1]:
        cell.font = _BOLD
    for d in result.differences:
        c, a = d["company"], d["anaf"]
        wd.append([d["diff_type"], d["line_no"], d["label"],
                   c["base"] if c else "", c["vat"] if c else "",
                   a["base"] if a else "", a["vat"] if a else "",
                   d["delta_base"], d["delta_vat"]])

    if facturi_suspecte or linii_neelucidate:
        wf = wb.create_sheet("Facturi de verificat", 0)
        wf.append(_SUSPECTE_HEADER)
        for cell in wf[1]:
            cell.font = _BOLD
        for f in (facturi_suspecte or []):
            wf.append([f.get("invoice_no", ""), f.get("date", ""),
                       f.get("partner_cui", ""), f.get("base"), f.get("vat"),
                       f.get("line_no", ""),
                       _MOTIV_SUSPECT.get(f.get("motiv"), f.get("motiv", ""))])
            if f.get("motiv") == "candidat":
                for cell in wf[wf.max_row]:
                    cell.fill = _RED
        if linii_neelucidate:
            wf.append([])
            wf.append(["Linii cu diferenta fara factura identificabila — "
                       "cauza nu e printre facturile din jurnal:"])
            wf[wf.max_row][0].font = _BOLD
            wf.append(_NEELUCIDATE_HEADER)
            for cell in wf[wf.max_row]:
                cell.font = _BOLD
            for l in linii_neelucidate:
                wf.append([l.get("line_no", ""), l.get("label", ""),
                           l.get("delta_base"), l.get("delta_vat"),
                           l.get("explicatie", "")])
    wb.save(path)
