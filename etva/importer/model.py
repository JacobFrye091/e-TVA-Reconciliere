"""Parser and generator for the e-TVA "model" journal format, offered as an
alternative to firms that don't export from SAGA: two downloadable .xlsx
templates (one for "vanzari", one for "cumparari") where the "Tip
operatiune" column is filled from a fixed dropdown wired directly onto a
D300 line number. Because the vocabulary is closed and chosen by the user
rather than free text, classification is deterministic — no heuristics
needed, unlike the SAGA legend (see etva.d300.suggest_line).
"""
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

from .saga import SagaJournal, _norm, _num, _cell, _date_str


class NotModelFormat(Exception):
    pass


MARKER = {
    "vanzari": "MODEL e-TVA - JURNAL DE VANZARI (v1)",
    "cumparari": "MODEL e-TVA - JURNAL DE CUMPARARI (v1)",
}

ANTET = ["Data", "Numar document", "Denumire partener", "Cod fiscal partener",
         "Baza impozitare", "Valoare TVA", "Tip operatiune"]

_COL_WIDTHS = [12, 16, 28, 18, 14, 14, 46]

HEADER_ROW_EXCEL = 4          # rand Excel (1-indexat) al antetului in foaia "Jurnal"
FIRST_DATA_ROW_EXCEL = HEADER_ROW_EXCEL + 1

# Etichetele de mai jos sunt cele afisate in dropdown-ul "Tip operatiune" si
# sunt formulate deliberat ca sa fie fara ambiguitate pentru
# etva.d300.suggest_line — fiecare trebuie sa produca fie exact linia D300
# asociata, fie None, niciodata alta linie (vezi garda din
# tests/test_model_jurnal.py). De-asta: "13" nu mentioneaza art. 331 (ar
# clasifica gresit pe 26.x), "20.1" nu mentioneaza o cota (ar cadea pe 24),
# "29" nu mentioneaza "intracomunitar" (s-ar suprapune cu 29.1).
TIPURI_OPERATIUNE = {
    "vanzari": [
        ("Livrari de bunuri si prestari de servicii taxabile cu cota 21%", "9"),
        ("Livrari de bunuri si prestari de servicii taxabile cu cota 11%", "10"),
        ("Livrari de bunuri si prestari de servicii taxabile cu cota 9%", "11"),
        ("Livrari intracomunitare de bunuri, scutite conform art. 294", "1"),
        ("Livrari supuse masurilor de simplificare (taxare inversa), cota 21%", "13"),
        ("Livrari/prestari cu locul in afara Romaniei (export, servicii UE)", "3"),
        ("Livrari scutite cu sau fara drept de deducere, in tara", "14+15"),
    ],
    "cumparari": [
        ("Achizitii de bunuri si servicii din tara, taxabile cu cota 21%", "24"),
        ("Achizitii de bunuri si servicii din tara, taxabile cu cota 11%", "25"),
        ("Achizitii intracomunitare de bunuri (AIC)", "20.1"),
        ("Achizitii de servicii intracomunitare, taxare inversa conform art. 307", "22.1"),
        ("Achizitii cu taxare inversa in tara conform art. 331, cota 21%", "26.1"),
        ("Achizitii cu taxare inversa in tara conform art. 331, cota 11%", "26.2"),
        ("Achizitii scutite de taxa sau neimpozabile", "29"),
    ],
}

_PASI_INSTRUCTIUNI = [
    "1. Adauga cate un rand pe fiecare factura, incepand cu randul "
    f"{FIRST_DATA_ROW_EXCEL} din foaia 'Jurnal'.",
    "2. Coloana 'Tip operatiune' se alege din lista derulanta — nu scrie text liber.",
    "3. Lasa liniile fara facturi goale; nu sterge si nu redenumi coloanele din antet.",
    "4. Nu include operatiuni cu TVA la incasare (neexigibila) — acestea se "
    "raporteaza separat, in perioada in care taxa devine exigibila.",
    "5. Salveaza fisierul ca .xlsx (nu .csv, nu .xls) inainte de a-l incarca in aplicatie.",
]


def build_model_template(directie: str) -> Workbook:
    tipuri = TIPURI_OPERATIUNE[directie]
    bold = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jurnal"

    ws["A1"] = MARKER[directie]
    ws["A1"].font = bold
    ws["A2"] = ("Completeaza cate un rand pe factura mai jos. Vezi foaia "
                "'Instructiuni' pentru detalii.")

    for i, titlu in enumerate(ANTET):
        cell = ws.cell(row=HEADER_ROW_EXCEL, column=i + 1, value=titlu)
        cell.font = bold

    for i, width in enumerate(_COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = width

    dv = DataValidation(type="list", formula1="=Liste!$A$1:$A$7", allow_blank=True)
    dv.error = "Alege o valoare din lista pentru 'Tip operatiune'."
    dv.errorTitle = "Tip operatiune invalid"
    dv.showErrorMessage = True
    ws.add_data_validation(dv)
    dv.add(f"G{FIRST_DATA_ROW_EXCEL}:G1000")

    ws_instr = wb.create_sheet("Instructiuni")
    ws_instr.column_dimensions["A"].width = 70
    ws_instr.column_dimensions["B"].width = 60
    ws_instr["A1"] = f"Cum completezi modelul e-TVA de {directie}"
    ws_instr["A1"].font = Font(bold=True, size=13)

    row = 3
    for text in _PASI_INSTRUCTIUNI:
        ws_instr.cell(row=row, column=1, value=text)
        row += 1

    row += 1
    ws_instr.cell(row=row, column=1, value="Tip operatiune").font = bold
    ws_instr.cell(row=row, column=2, value="Corespunde in formularul 300").font = bold
    row += 1
    for eticheta, linie in tipuri:
        ws_instr.cell(row=row, column=1, value=eticheta)
        ws_instr.cell(row=row, column=2, value=f"Randul {linie}")
        row += 1

    ws_liste = wb.create_sheet("Liste")
    for i, (eticheta, _linie) in enumerate(tipuri):
        ws_liste.cell(row=i + 1, column=1, value=eticheta)
    ws_liste.sheet_state = "hidden"

    return wb


def _read_raw(path: str) -> pd.DataFrame:
    engine = "xlrd" if path.lower().endswith(".xls") else None
    try:
        return pd.read_excel(path, header=None, engine=engine)
    except Exception as e:
        raise NotModelFormat(
            "Fisierul nu a putut fi citit ca Excel (.xls/.xlsx). Salveaza "
            "modelul e-TVA ca .xlsx si incarca-l din nou.") from e


def _detect_direction(df: pd.DataFrame) -> str:
    bases = {d: _norm(m).replace(" (v1)", "") for d, m in MARKER.items()}
    for r in range(min(5, len(df))):
        for c in range(df.shape[1]):
            text = _norm(_cell(df, r, c))
            if not text:
                continue
            for directie, base in bases.items():
                if base in text:
                    return directie
    raise NotModelFormat(
        "Fisierul nu contine marcajul 'MODEL e-TVA' asteptat pe primele "
        "randuri. Descarca sablonul din aplicatie (optiunea 'Alt program "
        "de contabilitate') si completeaza-l fara sa modifici antetul.")


_COLUMN_KEYWORDS = {
    "date": ("data",),
    "doc_no": ("numar document", "numar"),
    "partner_name": ("denumire partener", "denumire"),
    "partner_cui": ("cod fiscal partener", "cod fiscal"),
    "base": ("baza impozitare", "baza"),
    "vat": ("valoare tva", "valoare t.v.a."),
    "tip": ("tip operatiune",),
}


def _find_header_row(df: pd.DataFrame) -> int:
    for r in range(min(10, len(df))):
        for c in range(df.shape[1]):
            if _norm(_cell(df, r, c)) == "tip operatiune":
                return r
    raise NotModelFormat(
        "Nu s-a gasit antetul coloanelor (randul cu 'Tip operatiune'). "
        "Descarca din nou modelul din aplicatie si nu modifica antetul.")


def _find_columns(df: pd.DataFrame, header_row: int) -> dict:
    found = {}
    for c in range(df.shape[1]):
        text = _norm(_cell(df, header_row, c))
        for field_name, keywords in _COLUMN_KEYWORDS.items():
            if field_name in found:
                continue
            if any(kw in text for kw in keywords):
                found[field_name] = c
    missing = [f for f in _COLUMN_KEYWORDS if f not in found]
    if missing:
        raise NotModelFormat(
            f"Coloane lipsa in modelul e-TVA: {', '.join(missing)}. "
            "Descarca din nou modelul din aplicatie si nu modifica antetul.")
    return found


def parse_model_journal(path: str) -> SagaJournal:
    df = _read_raw(path)
    direction = _detect_direction(df)
    header_row = _find_header_row(df)
    cols = _find_columns(df, header_row)
    label_to_line = {_norm(eticheta): linie
                     for eticheta, linie in TIPURI_OPERATIUNE[direction]}

    entries = []
    legend: dict = {}
    for r in range(header_row + 1, len(df)):
        cui = _cell(df, r, cols["partner_cui"])
        doc_no = _cell(df, r, cols["doc_no"])
        base = _cell(df, r, cols["base"])
        vat = _cell(df, r, cols["vat"])
        tip = _cell(df, r, cols["tip"])
        if cui is None and doc_no is None and base is None and vat is None and tip is None:
            continue  # rand gol, lasat neatins din sablon
        if tip is None or not str(tip).strip():
            raise NotModelFormat(
                f"Randul {r + 1}: lipseste 'Tip operatiune' — alege o "
                "valoare din lista derulanta.")
        try:
            base_v = _num(base)
            vat_v = _num(vat)
        except ValueError as e:
            raise NotModelFormat(
                f"Randul {r + 1}: 'Baza impozitare' sau 'Valoare TVA' nu "
                "sunt numere valide.") from e

        label = str(tip).strip()
        cod = label_to_line.get(_norm(label), label)
        entries.append({
            "date": _date_str(_cell(df, r, cols["date"])),
            "doc_no": str(doc_no or "").strip(),
            "partner_name": str(_cell(df, r, cols["partner_name"]) or "").strip(),
            "partner_cui": str(cui or "").strip().upper(),
            "base": base_v,
            "vat": vat_v,
            "cod": cod,
        })
        acc = legend.setdefault(cod, {"label": label, "base": 0.0, "vat": 0.0})
        acc["base"] += base_v
        acc["vat"] += vat_v

    return SagaJournal(direction=direction, company_name=None, company_cui=None,
                       entries=entries, legend=legend)
