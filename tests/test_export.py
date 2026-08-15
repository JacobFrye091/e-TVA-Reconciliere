from openpyxl import load_workbook
from etva.engine import reconcile
from etva.advisor import suggest_d300
from etva import export

def row(cui="RO1", no="F1", base=100.0, vat=19.0, cat="livrari_interne"):
    return {"partner_cui": cui, "invoice_no": no, "date": "2026-01-10",
            "base": base, "vat": vat, "category": cat}

def test_report_structure(tmp_path):
    r = reconcile([row()], [row(base=200.0)])
    p = str(tmp_path / "raport.xlsx")
    export.write_report(r, suggest_d300(r), p, "Firma SRL", "2026-01")
    wb = load_workbook(p)
    assert wb.sheetnames == ["Sumar", "Diferente"]
    sumar = wb["Sumar"]
    assert sumar["A1"].value == "Client: Firma SRL"
    assert sumar["A2"].value == "Perioada: 2026-01"
    diffs = wb["Diferente"]
    assert diffs["A1"].value == "Tip diferenta"
    assert diffs["A2"].value == "suma_diferita"

def test_flagged_row_is_red(tmp_path):
    r = reconcile([row()], [row(base=200.0)])
    p = str(tmp_path / "raport.xlsx")
    export.write_report(r, suggest_d300(r), p, "F", "2026-01")
    sumar = load_workbook(p)["Sumar"]
    # data starts at row 5 (title, period, blank, header)
    assert sumar.cell(row=5, column=1).fill.start_color.rgb == "00FFC7CE"


def _rezultat_linii():
    from etva.engine import reconcile_d300
    return reconcile_d300({"9": {"base": 160.5, "vat": 33.71}},
                          {"9": {"base": 100.0, "vat": 21.0}})


def test_raport_pe_linii_fara_facturi_suspecte_nu_adauga_foaia(tmp_path):
    """Compatibilitate: apelantii vechi (fara argumentele noi) primesc
    exact acelasi fisier ca inainte."""
    from etva.advisor import suggest_d300_lines
    r = _rezultat_linii()
    p = str(tmp_path / "raport.xlsx")
    export.write_report_lines(r, suggest_d300_lines(r), p, "Firma SRL", "2026-06")
    assert load_workbook(p).sheetnames == ["Sumar", "Diferente"]


def test_raport_pe_linii_include_foaia_facturi_de_verificat(tmp_path):
    """Foaia e prima in fisier: raspunde la intrebarea cu care incepe orice
    verificare - de unde incep."""
    from etva.advisor import suggest_d300_lines
    r = _rezultat_linii()
    p = str(tmp_path / "raport.xlsx")
    export.write_report_lines(
        r, suggest_d300_lines(r), p, "Firma SRL", "2026-06",
        facturi_suspecte=[{"invoice_no": "F2", "date": "2026-06-02",
                           "partner_cui": "RO222", "base": 60.5, "vat": 12.71,
                           "line_no": "9", "motiv": "candidat"}],
        linii_neelucidate=[{"line_no": "24", "label": "Achizitii",
                            "delta_base": -500.0, "delta_vat": -105.0,
                            "explicatie": "ANAF are inregistrat mai mult."}])
    wb = load_workbook(p)
    assert wb.sheetnames[0] == "Facturi de verificat"
    wf = wb["Facturi de verificat"]
    assert wf["A1"].value == "Document"
    assert wf["A2"].value == "F2"
    assert wf["F2"].value == "9"
    assert "Posibila cauza" in wf["G2"].value
    # Randul de factura-cauza e evidentiat, ca in ecran.
    assert wf.cell(row=2, column=1).fill.start_color.rgb == "00FFC7CE"
    # Linia neelucidata apare undeva mai jos, cu explicatia ei.
    text = "\n".join(str(c.value) for r_ in wf.iter_rows() for c in r_ if c.value)
    assert "24" in text and "ANAF are inregistrat mai mult." in text
